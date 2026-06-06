# ============================================================
#  Student Result Portal — app.py
#  WhatsApp notifications via Meta Cloud API (FREE)
#  Place this file in: student_result_portal/app.py
# ============================================================

from flask import (
    Flask, render_template, request,
    redirect, url_for, session, jsonify, flash
)
import sqlite3, os, hashlib

try:
    import requests as http_req
except ImportError:
    os.system("pip install requests")
    import requests as http_req

app = Flask(__name__)
app.secret_key = "srp_ultra_secret_2025"
DB_PATH = os.path.join(os.path.dirname(__file__), "students.db")


# ════════════════════════════════════════════════════════════════
#  META WHATSAPP CLOUD API CREDENTIALS
#  Full setup guide at the bottom of this file (read the comments)
# ════════════════════════════════════════════════════════════════
META_ACCESS_TOKEN  = "EAAjjC8wxc10BQ5XgqUrhq2rSQgarLQyHIVEsao7rMJpBmDm70JO6IR51n98QYrJk7acUI9QATwVixJYgGDAe3NpbzrJR4ZA4EWwqka3kVIrFIY1DGDKvnqbQ6db4gvWokUkrhqrgxHb1Wh1IRd8nz27roBRK5tCAXpZCOL2hrwRg3T6zQzNcEZA9iLdjde4WNZAsVxV35BqtUwIZB9X6KaQTmx7e6LPJZALSQMhm7a4FWdEiAVGM9MMPd7klC0DgiD0Wj7cfagpAwF8OzeSQc5ftsM"    # paste from Meta dashboard
META_PHONE_NUM_ID  = "YO972720099263987"      # paste from Meta dashboard
META_API_VERSION   = "v21.0"
META_TEMPLATE_NAME = "result_notification"       # must match your approved template name
META_TEMPLATE_LANG = "en"
META_API_URL = (
    f"https://graph.facebook.com/{META_API_VERSION}"
    f"/{META_PHONE_NUM_ID}/messages"
)

# Admin panel password — change this!
ADMIN_PASSWORD = "admin1234"

# Subject max-marks
SUBJECT_TOTAL = {
    "Mathematics": 200, "English": 200, "Science": 200,
    "Urdu": 150, "Computer": 150, "Islamiyat": 100,
}


# ════════════════════════════════════════════════════════════════
#  UTILITIES
# ════════════════════════════════════════════════════════════════

def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ════════════════════════════════════════════════════════════════
#  META WHATSAPP SEND FUNCTION
# ════════════════════════════════════════════════════════════════

def send_whatsapp_meta(phone_number: str, student_name: str,
                       obtained: int, total: int,
                       grade: str, percentage: float) -> dict:
    """
    Send WhatsApp template message via Meta Cloud API.
    Template must have 5 parameters: name, obtained, total, grade, percentage
    Returns {"success": True/False, "info": "..."}
    """
    # Normalise to E.164 (no + prefix for Meta)
    phone = str(phone_number).strip().replace(" ", "").replace("-", "")
    if phone.startswith("0"):
        phone = phone[1:]
    if not phone.startswith("92"):
        phone = "92" + phone
    if phone.startswith("+"):
        phone = phone[1:]

    headers = {
        "Authorization": f"Bearer {META_ACCESS_TOKEN}",
        "Content-Type":  "application/json",
    }
    payload = {
        "messaging_product": "whatsapp",
        "to":   phone,
        "type": "template",
        "template": {
            "name":     META_TEMPLATE_NAME,
            "language": {"code": META_TEMPLATE_LANG},
            "components": [{
                "type": "body",
                "parameters": [
                    {"type": "text", "text": student_name},
                    {"type": "text", "text": str(obtained)},
                    {"type": "text", "text": str(total)},
                    {"type": "text", "text": grade},
                    {"type": "text", "text": str(percentage)},
                ]
            }]
        }
    }

    try:
        resp = http_req.post(META_API_URL, headers=headers, json=payload, timeout=15)
        data = resp.json()

        if resp.status_code == 200 and "messages" in data:
            return {"success": True, "info": f"Sent ✓  ID: {data['messages'][0].get('id','')}"}

        error = data.get("error", {})
        code  = error.get("code", resp.status_code)
        msg   = error.get("message", str(data))
        hints = {
            190:    "Access token expired — regenerate in Meta dashboard.",
            100:    "Invalid parameter — check Phone Number ID or template name.",
            131030: "Template not found or not yet approved — check WhatsApp Manager.",
            131047: "Not a test recipient — add number in Meta dashboard first.",
            131026: "This number has no WhatsApp account.",
            368:    "Account restricted — check Meta Business account.",
        }
        note = hints.get(code, "")
        return {"success": False, "info": f"Error {code}: {msg}" + (f" → {note}" if note else "")}

    except http_req.exceptions.Timeout:
        return {"success": False, "info": "Request timed out."}
    except Exception as e:
        return {"success": False, "info": str(e)}


# ════════════════════════════════════════════════════════════════
#  AI ANALYSIS ENGINE
# ════════════════════════════════════════════════════════════════

def analyse_subjects(subject_marks: dict) -> dict:
    TIPS = {
        "Mathematics": ["Practice 10 problems daily.", "Build a formula sheet.", "Solve 5 years of past papers."],
        "English":     ["Read one article daily.", "Write a paragraph daily.", "Focus on grammar: tenses and articles."],
        "Science":     ["Sketch diagrams by hand.", "Connect theory to real life.", "Revise formulas every Sunday."],
        "Urdu":        ["Read Urdu newspaper 15 min daily.", "Write Urdu essays on exam topics.", "Keep a vocabulary notebook."],
        "Computer":    ["Practice typing every session.", "Highlight key terms while reading.", "Do hands-on tasks like Word/Excel."],
        "Islamiyat":   ["Memorise key verses with translation.", "Learn important Islamic history dates.", "Write concise topic notes."],
    }
    STATUS_MAP = [
        (90,"Outstanding","#22d3a5"),(80,"Excellent","#3ecf7a"),
        (70,"Good","#7aadff"),(55,"Average","#d4a843"),
        (40,"Weak","#f09060"),(0,"Critical","#f06060"),
    ]
    results={}; weak=[]; strong=[]; total_pct=0.0
    for subj, obtained in subject_marks.items():
        max_m = SUBJECT_TOTAL.get(subj, 100)
        pct   = round((obtained/max_m)*100, 1)
        total_pct += pct
        status,color = "Average","#d4a843"
        for thr,lbl,clr in STATUS_MAP:
            if pct >= thr:
                status,color = lbl,clr; break
        results[subj] = {
            "obtained":obtained,"total":max_m,"pct":pct,
            "status":status,"color":color,
            "tips":TIPS.get(subj,[]) if status in ("Weak","Critical") else [],
        }
        if status in ("Weak","Critical"): weak.append(subj)
        elif status in ("Outstanding","Excellent","Good"): strong.append(subj)
    avg_pct = round(total_pct/len(subject_marks),1) if subject_marks else 0
    if not weak:
        advice = "Exceptional work! Maintain discipline and explore advanced material."
    elif len(weak)==1:
        advice = f"Great overall! Focus extra 30 min daily on {weak[0]} to improve fast."
    elif len(weak)==2:
        advice = f"{weak[0]} and {weak[1]} need immediate attention with dedicated daily slots."
    else:
        advice = "Several subjects need work. Build a rotating timetable and tackle one subject per day."
    return {"subjects":results,"weak":weak,"strong":strong,"avg_pct":avg_pct,"overall_advice":advice}


# ════════════════════════════════════════════════════════════════
#  EXCEL DATA LOADER
# ════════════════════════════════════════════════════════════════

EXCEL_FILE     = os.path.join(os.path.dirname(__file__), "student_data.xlsx")
VALID_SUBJECTS = set(SUBJECT_TOTAL.keys())

def load_from_excel():
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl"); return [],[]
    if not os.path.exists(EXCEL_FILE):
        print(f"'{EXCEL_FILE}' not found."); return [],[]
    print(f"Loading: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    students=[]; marks=[]
    if "Students" in wb.sheetnames:
        for row in wb["Students"].iter_rows(min_row=5, values_only=True):
            if not row[0]: continue
            sid  = str(row[0]).strip().upper()
            name = str(row[1]).strip() if row[1] else ""
            cls  = str(row[2]).strip() if row[2] else ""
            total=row[3]; obtained=row[4]
            grade= str(row[5]).strip() if row[5] else ""
            pw   = str(row[6]).strip() if row[6] else sid.lower()
            phone= str(row[7]).strip() if len(row)>7 and row[7] else ""
            if not all([sid,name,cls,total,obtained,grade]): continue
            try: students.append((sid,name,cls,int(total),int(obtained),grade,pw,phone))
            except: continue
    if "Subject Marks" in wb.sheetnames:
        for row in wb["Subject Marks"].iter_rows(min_row=5, values_only=True):
            if not row[0]: continue
            sid=str(row[0]).strip().upper()
            subj=str(row[1]).strip() if row[1] else ""
            if subj not in VALID_SUBJECTS: continue
            try: marks.append((sid,subj,int(row[2])))
            except: continue
    print(f"  ✓ {len(students)} students | {len(marks)} marks")
    return students, marks


# ════════════════════════════════════════════════════════════════
#  DATABASE INIT
# ════════════════════════════════════════════════════════════════

def init_db():
    conn=get_db(); cur=conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL, email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL, role TEXT DEFAULT 'student')""")
    cur.execute("""CREATE TABLE IF NOT EXISTS students (
        id TEXT PRIMARY KEY, name TEXT NOT NULL, class TEXT NOT NULL,
        total INTEGER NOT NULL, obtained INTEGER NOT NULL, grade TEXT NOT NULL,
        parent_phone TEXT DEFAULT '', whatsapp_sent INTEGER DEFAULT 0)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS subject_marks (
        student_id TEXT NOT NULL, subject TEXT NOT NULL, obtained INTEGER NOT NULL,
        PRIMARY KEY (student_id, subject),
        FOREIGN KEY (student_id) REFERENCES students(id))""")
    for col,defn in [("parent_phone","TEXT DEFAULT ''"),("whatsapp_sent","INTEGER DEFAULT 0")]:
        try: cur.execute(f"ALTER TABLE students ADD COLUMN {col} {defn}")
        except: pass

    students, marks = load_from_excel()
    for sid,name,cls,total,obtained,grade,pw,phone in students:
        cur.execute("""INSERT INTO students (id,name,class,total,obtained,grade,parent_phone)
            VALUES (?,?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET
            name=excluded.name,class=excluded.class,total=excluded.total,
            obtained=excluded.obtained,grade=excluded.grade,parent_phone=excluded.parent_phone
        """, (sid,name,cls,total,obtained,grade,phone))
        cur.execute("""INSERT INTO users (name,email,password) VALUES (?,?,?)
            ON CONFLICT(email) DO UPDATE SET name=excluded.name,password=excluded.password
        """, (name,sid,hash_pw(pw)))
    for sid,subj,obt in marks:
        cur.execute("""INSERT INTO subject_marks (student_id,subject,obtained) VALUES (?,?,?)
            ON CONFLICT(student_id,subject) DO UPDATE SET obtained=excluded.obtained
        """, (sid,subj,obt))
    conn.commit(); conn.close()


# ════════════════════════════════════════════════════════════════
#  DECORATORS
# ════════════════════════════════════════════════════════════════

def login_required(f):
    from functools import wraps
    @wraps(f)
    def d(*a,**kw):
        if "student" not in session: return redirect(url_for("login"))
        return f(*a,**kw)
    return d

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def d(*a,**kw):
        if not session.get("is_admin"): return redirect(url_for("admin_login"))
        return f(*a,**kw)
    return d

def fetch_student_data(sid):
    conn=get_db()
    row=conn.execute("SELECT * FROM students WHERE id=?",(sid,)).fetchone()
    sr =conn.execute("SELECT subject,obtained FROM subject_marks WHERE student_id=?",(sid,)).fetchall()
    conn.close()
    if not row: return None,None
    pct=round((row["obtained"]/row["total"])*100,2)
    student={"id":row["id"],"name":row["name"],"cls":row["class"],
             "total":row["total"],"obtained":row["obtained"],"grade":row["grade"],
             "percentage":pct,"parent_phone":row["parent_phone"] or "",
             "whatsapp_sent":row["whatsapp_sent"]}
    sm={r["subject"]:r["obtained"] for r in sr}
    return student,(analyse_subjects(sm) if sm else None)


# ════════════════════════════════════════════════════════════════
#  STUDENT ROUTES
# ════════════════════════════════════════════════════════════════

@app.route("/",methods=["GET"])
@app.route("/login",methods=["GET","POST"])
def login():
    if "student" in session: return redirect(url_for("dashboard"))
    error=None
    if request.method=="POST":
        sid=request.form.get("student_id","").strip().upper()
        pw =request.form.get("password","").strip()
        if not sid or not pw:
            error="Both Student ID and Password are required."
        else:
            conn=get_db()
            user=conn.execute("SELECT * FROM users WHERE email=? AND password=?",
                              (sid,hash_pw(pw))).fetchone()
            conn.close()
            if user:
                session["student"]={"id":sid,"name":user["name"]}
                return redirect(url_for("dashboard"))
            error="Invalid Student ID or Password."
    return render_template("login.html",error=error)

@app.route("/logout")
def logout():
    session.pop("student",None); return redirect(url_for("login"))

@app.route("/dashboard")
@login_required
def dashboard():
    sid=session["student"]["id"]
    student,ai=fetch_student_data(sid)
    if not student: session.clear(); return redirect(url_for("login"))
    return render_template("dashboard.html",student=student,ai=ai,user=session["student"])

@app.route("/ai-analysis")
@login_required
def ai_analysis():
    sid=session["student"]["id"]
    student,ai=fetch_student_data(sid)
    if not student: session.clear(); return redirect(url_for("login"))
    return render_template("ai_analysis.html",student=student,ai=ai,user=session["student"])

@app.route("/api/chart-data")
@login_required
def chart_data():
    sid=session["student"]["id"]
    conn=get_db()
    rows=conn.execute("SELECT subject,obtained FROM subject_marks WHERE student_id=?",(sid,)).fetchall()
    conn.close()
    labels=[r["subject"] for r in rows]; values=[r["obtained"] for r in rows]
    totals=[SUBJECT_TOTAL.get(r["subject"],100) for r in rows]
    pcts=[round(v/t*100,1) for v,t in zip(values,totals)]
    return jsonify({"labels":labels,"values":values,"totals":totals,"percentages":pcts})


# ════════════════════════════════════════════════════════════════
#  ADMIN ROUTES
# ════════════════════════════════════════════════════════════════

@app.route("/admin/login",methods=["GET","POST"])
def admin_login():
    error=None
    if request.method=="POST":
        if request.form.get("password","").strip()==ADMIN_PASSWORD:
            session["is_admin"]=True; return redirect(url_for("admin_panel"))
        error="Wrong admin password."
    return render_template("admin_login.html",error=error)

@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin",None); return redirect(url_for("admin_login"))

@app.route("/admin")
@admin_required
def admin_panel():
    conn=get_db()
    students=conn.execute(
        "SELECT id,name,class,obtained,total,grade,parent_phone,whatsapp_sent "
        "FROM students ORDER BY class,name").fetchall()
    conn.close()
    creds_ok=(META_ACCESS_TOKEN!="YOUR_META_ACCESS_TOKEN"
              and META_PHONE_NUM_ID!="YOUR_PHONE_NUMBER_ID")
    return render_template("admin_panel.html",
        students=students,
        total_students=len(students),
        with_phone=sum(1 for s in students if s["parent_phone"]),
        wa_sent_count=sum(1 for s in students if s["whatsapp_sent"]),
        credentials_set=creds_ok,
        template_name=META_TEMPLATE_NAME,
    )

@app.route("/admin/send-all",methods=["POST"])
@admin_required
def admin_send_all():
    conn=get_db()
    students=conn.execute(
        "SELECT id,name,obtained,total,grade,parent_phone FROM students "
        "WHERE parent_phone!='' AND parent_phone IS NOT NULL").fetchall()
    skipped=conn.execute(
        "SELECT COUNT(*) FROM students WHERE parent_phone='' OR parent_phone IS NULL"
    ).fetchone()[0]
    conn.close()
    sent=0; failed=0; results=[]
    for s in students:
        pct=round((s["obtained"]/s["total"])*100,1)
        result=send_whatsapp_meta(s["parent_phone"],s["name"],
                                  s["obtained"],s["total"],s["grade"],pct)
        if result["success"]:
            sent+=1
            conn=get_db()
            conn.execute("UPDATE students SET whatsapp_sent=1 WHERE id=?",(s["id"],))
            conn.commit(); conn.close()
        else:
            failed+=1
        results.append({"id":s["id"],"name":s["name"],"phone":s["parent_phone"],
                         "success":result["success"],"info":result["info"]})
    return render_template("admin_wa_result.html",
                           results=results,sent=sent,failed=failed,skipped=skipped)

@app.route("/admin/send-one/<student_id>",methods=["POST"])
@admin_required
def admin_send_one(student_id):
    conn=get_db()
    s=conn.execute("SELECT * FROM students WHERE id=?",(student_id,)).fetchone()
    conn.close()
    if not s or not s["parent_phone"]:
        flash(f"No parent phone for {student_id}.","error")
        return redirect(url_for("admin_panel"))
    pct=round((s["obtained"]/s["total"])*100,1)
    result=send_whatsapp_meta(s["parent_phone"],s["name"],
                               s["obtained"],s["total"],s["grade"],pct)
    if result["success"]:
        conn=get_db()
        conn.execute("UPDATE students SET whatsapp_sent=1 WHERE id=?",(student_id,))
        conn.commit(); conn.close()
        flash(f"✅ WhatsApp sent to parent of {s['name']}!","success")
    else:
        flash(f"❌ Failed for {s['name']}: {result['info']}","error")
    return redirect(url_for("admin_panel"))

@app.route("/admin/test-whatsapp",methods=["POST"])
@admin_required
def admin_test_whatsapp():
    test_phone=request.form.get("test_phone","").strip()
    if not test_phone:
        flash("Enter a phone number to test.","error")
        return redirect(url_for("admin_panel"))
    result=send_whatsapp_meta(test_phone,"Test Student",850,1000,"A",85.0)
    if result["success"]:
        flash(f"✅ Test WhatsApp sent to {test_phone}! Check the phone.","success")
    else:
        flash(f"❌ Test failed: {result['info']}","error")
    return redirect(url_for("admin_panel"))

@app.route("/admin/reset-whatsapp",methods=["POST"])
@admin_required
def admin_reset_whatsapp():
    conn=get_db()
    conn.execute("UPDATE students SET whatsapp_sent=0")
    conn.commit(); conn.close()
    flash("WhatsApp sent flags reset. You can now resend to all parents.","info")
    return redirect(url_for("admin_panel"))


# ════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════════

if __name__=="__main__":
    init_db()
    app.run(debug=True)


# ════════════════════════════════════════════════════════════════
#  META WHATSAPP SETUP GUIDE (read once, then delete this section)
# ════════════════════════════════════════════════════════════════
#
#  STEP 1 — Create Meta Developer App (15 min)
#    1. Go to: developers.facebook.com
#    2. Click "My Apps" → "Create App" → choose "Business"
#    3. Add "WhatsApp" product to the app
#    4. You instantly get a TEST phone number and a temporary token
#
#  STEP 2 — Get your credentials
#    - ACCESS_TOKEN  : App Dashboard → WhatsApp → API Setup → copy the token
#    - PHONE_NUM_ID  : Same page → copy the "Phone number ID" shown
#    - Paste both into META_ACCESS_TOKEN and META_PHONE_NUM_ID above
#
#  STEP 3 — Create your message template
#    1. Go to: business.facebook.com → WhatsApp Manager → Message Templates
#    2. Click "Create Template"
#    3. Category: Utility  |  Name: result_notification  |  Language: English
#    4. Template body (copy exactly):
#
#         Assalam o Alaikum!
#         Result of {{1}}:
#         Marks: {{2}}/{{3}}
#         Grade: {{4}}
#         Percentage: {{5}}%
#         See full result at: your-school-website.com
#
#    5. Submit for review — Meta approves in 24-48 hours
#
#  STEP 4 — For testing BEFORE approval
#    - In Meta dashboard → WhatsApp → API Setup → "To" field
#    - Add your own phone number as a "test recipient"
#    - You can send to test recipients immediately without template approval
#
#  STEP 5 — Add parent phones to Excel
#    - Open student_data.xlsx → Students sheet → Column H = "Parent Phone"
#    - Format: 03001234567 (with or without leading 0, app handles it)
#    - Save file, restart python app.py
#
#  ADMIN PANEL URL: http://127.0.0.1:5000/admin/login
#  Default admin password: admin1234  (change ADMIN_PASSWORD above!)
#
# ════════════════════════════════════════════════════════════════
