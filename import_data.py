# ============================================================
#  import_data.py
#  Place this file in: student_result_portal/import_data.py
#
#  HOW TO USE:
#    1. Fill in student_data.xlsx (Students + Subject Marks sheets)
#    2. Open terminal in your project folder
#    3. Run:  python import_data.py
#    4. Done! All students are loaded into students.db
# ============================================================

import sqlite3
import hashlib
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

# ── Config ────────────────────────────────────────────────────
EXCEL_FILE = "student_data.xlsx"   # Name of your Excel file
DB_FILE    = "students.db"         # Your database file

# Subject max marks (must match what's in your portal)
SUBJECT_MAX = {
    "Mathematics": 200,
    "English":     200,
    "Science":     200,
    "Urdu":        150,
    "Computer":    150,
    "Islamiyat":   100,
}

# ── Helpers ───────────────────────────────────────────────────
def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def clean(val):
    """Strip whitespace from string values."""
    return str(val).strip() if val is not None else ""

# ── Main ──────────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  AI Result Portal — Data Import Tool")
    print("=" * 55)

    # Check Excel file exists
    if not os.path.exists(EXCEL_FILE):
        print(f"\n❌  ERROR: '{EXCEL_FILE}' not found!")
        print(f"   Make sure '{EXCEL_FILE}' is in the same folder as this script.")
        sys.exit(1)

    print(f"\n📂  Reading:  {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # ── Read Sheet 1: Students ────────────────────────────────
    if "Students" not in wb.sheetnames:
        print("❌  ERROR: 'Students' sheet not found in Excel file!")
        sys.exit(1)

    ws1 = wb["Students"]
    students = []
    errors   = []

    print("\n📋  Reading Students sheet...")

    for row in ws1.iter_rows(min_row=5, values_only=True):
        # Skip empty rows
        if not row[0]:
            continue

        sid      = clean(row[0]).upper()
        name     = clean(row[1])
        cls      = clean(row[2])
        total    = row[3]
        obtained = row[4]
        grade    = clean(row[5])
        password = clean(row[6]) if row[6] else sid.lower()

        # Validate
        if not all([sid, name, cls, total, obtained, grade]):
            errors.append(f"   Row skipped (missing data): {row}")
            continue

        try:
            total    = int(total)
            obtained = int(obtained)
        except (ValueError, TypeError):
            errors.append(f"   Row skipped (invalid marks): {sid}")
            continue

        students.append((sid, name, cls, total, obtained, grade, password))
        print(f"   ✓  {sid}  {name}  ({cls})  →  {obtained}/{total}  Grade: {grade}")

    print(f"\n   Found {len(students)} students.")

    # ── Read Sheet 2: Subject Marks ───────────────────────────
    if "Subject Marks" not in wb.sheetnames:
        print("❌  ERROR: 'Subject Marks' sheet not found in Excel file!")
        sys.exit(1)

    ws2 = wb["Subject Marks"]
    subject_marks = []

    print("\n📚  Reading Subject Marks sheet...")

    for row in ws2.iter_rows(min_row=5, values_only=True):
        if not row[0]:
            continue

        sid     = clean(row[0]).upper()
        subject = clean(row[1])
        try:
            obtained = int(row[2])
        except (ValueError, TypeError):
            errors.append(f"   Row skipped (invalid mark): {sid} - {subject}")
            continue

        if subject not in SUBJECT_MAX:
            errors.append(f"   Unknown subject '{subject}' for {sid} — skipped.")
            continue

        subject_marks.append((sid, subject, obtained))

    print(f"   Found {len(subject_marks)} subject mark entries.")

    # ── Connect to Database ───────────────────────────────────
    print(f"\n💾  Writing to database: {DB_FILE}")
    conn = sqlite3.connect(DB_FILE)
    cur  = conn.cursor()

    # Create tables if they don't exist
    cur.execute("""
        CREATE TABLE IF NOT EXISTS students (
            id TEXT PRIMARY KEY, name TEXT, class TEXT,
            total INTEGER, obtained INTEGER, grade TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS subject_marks (
            student_id TEXT, subject TEXT, obtained INTEGER,
            PRIMARY KEY (student_id, subject)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT, email TEXT UNIQUE, password TEXT, role TEXT DEFAULT 'student'
        )
    """)

    # ── Insert Students ───────────────────────────────────────
    inserted_s = 0
    updated_s  = 0
    for sid, name, cls, total, obtained, grade, password in students:
        # INSERT OR REPLACE so re-running is safe
        cur.execute("""
            INSERT INTO students (id, name, class, total, obtained, grade)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                name=excluded.name, class=excluded.class,
                total=excluded.total, obtained=excluded.obtained,
                grade=excluded.grade
        """, (sid, name, cls, total, obtained, grade))

        # Users table (login)
        cur.execute("""
            INSERT INTO users (name, email, password)
            VALUES (?, ?, ?)
            ON CONFLICT(email) DO UPDATE SET
                name=excluded.name, password=excluded.password
        """, (name, sid, hash_pw(password)))

        if cur.rowcount:
            inserted_s += 1

    # ── Insert Subject Marks ──────────────────────────────────
    inserted_m = 0
    for sid, subject, obtained in subject_marks:
        cur.execute("""
            INSERT INTO subject_marks (student_id, subject, obtained)
            VALUES (?, ?, ?)
            ON CONFLICT(student_id, subject) DO UPDATE SET
                obtained=excluded.obtained
        """, (sid, subject, obtained))
        inserted_m += 1

    conn.commit()
    conn.close()

    # ── Summary ───────────────────────────────────────────────
    print("\n" + "=" * 55)
    print("  ✅  IMPORT COMPLETE!")
    print("=" * 55)
    print(f"  Students loaded     : {len(students)}")
    print(f"  Subject rows loaded : {len(subject_marks)}")
    print(f"  Database file       : {DB_FILE}")

    if errors:
        print(f"\n  ⚠️  Warnings ({len(errors)}):")
        for e in errors:
            print(e)

    print("\n  ▶  Now run:  python app.py")
    print("  ▶  Open:     http://127.0.0.1:5000")
    print("=" * 55)

if __name__ == "__main__":
    main()
